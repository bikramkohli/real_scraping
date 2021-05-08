# Bikram Kohli (bsk4uaa@virginia.edu)
# Web scraping to find information on mobile home parks in a given state (location, distance, link to park, etc.)

'''
Importing selenium to access Google Chrome and access html code
Importing BeautifulSoup to parse through HTML code
Importing time to periodically pause code to allow webpages to load
Importing re to use regular expressions to check for valid addresses
Importing googlemaps and datetime to utilize Google Maps API to calculate trip duration to mobile home park
Importing openpyxl to write data to Excel file
'''

from selenium import webdriver
from bs4 import BeautifulSoup
import time
import re
import googlemaps
from datetime import datetime
from openpyxl import *


# regex matching helper function
def match_test(regex, text):
    # Gives a list of all complete matches
    ans = ''
    for match in regex.finditer(text):
        ans += match.group(0)
    return ans


# helper function to get trip duration to mobile home park
def getDuration(address):
    gmaps = googlemaps.Client(key='AIzaSyDXuFSMcQvP9pQ7wZwf-Oz3X5I6MFvelfQ')
    now = datetime.now()
    # finding duration
    directions_result = gmaps.directions("42559 Unbridleds Song Pl, South Riding, VA",
                                         address,
                                         mode="driving",
                                         departure_time=now)
    # getting specific duration time using Regular Expression and returning it
    directionsRE = re.compile(r'duration\':[a-z\'\{\:]+')
    return directions_result[0].get('legs')[0].get('duration').get('text')


# defining site URL
siteFirst = "https://www.mhvillage.com/parks/"
siteState = input("Enter a state abbreviation: ")
siteTotal = siteFirst + siteState + "?group-by=counties"

# opening site URL
chromedriver = "/Users/bikramkohli/Documents/chromedriver"
driver = webdriver.Chrome(chromedriver)
driver.get(siteTotal)
time.sleep(1 / 4)

# 'soup' object to be parsed
soup = BeautifulSoup(driver.page_source, "html.parser")

# finds all listed counties and stores them in 'counties' variable
time.sleep(1 / 4)
counties = soup.find_all('strong', {'class': "location-county"})

# defining iterator to write to specific Excel cells
i = 2

# parsing through each county
for county in counties:
    countyStr = county.text

    # opening up county
    driver.get(siteFirst + siteState + "/" + countyStr + "-county")
    time.sleep(1 / 4)

    # new 'soup' object for new site
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # finds all listed mobile home parks and stores them in 'parks' variable
    parks = soup.find_all('a', {'class': "text-decoration-none fc-brand-mhv-blue"})

    # parsing through each park
    for park in parks:
        try:
            parkName = park.text

            # regex matcher to find site for park
            parkHREFRE = re.compile(r'[/0-9]+["]')
            parkURL = siteFirst + match_test(parkHREFRE, str(park))[1:-1]

            # opening park
            driver.get(parkURL)
            time.sleep(2)
            soup = BeautifulSoup(driver.page_source, "html.parser")

            # getting address for parks
            streetAddress = soup.find('street-address-widget').text.strip()
            cityStateZip = soup.find('city-state-zip-widget').text.strip()
            streetAndCityStr = streetAddress + ", " + cityStateZip
            time.sleep(1 / 4)

            # opening up Excel file and printing data to file
            filepath = "/Users/bikramkohli/Documents/SWE Projects/mobileParksDataCounty.xlsx"
            wb = load_workbook(filepath)
            ws = wb["Sheet1"]

            # adding state
            wcell1 = ws.cell(i, 1)
            wcell1.value = siteState.upper()

            # adding city
            wcell2 = ws.cell(i, 2)
            city = cityStateZip.split(',')[0]
            wcell2.value = city

            # adding county
            wcell3 = ws.cell(i, 3)
            wcell3.value = countyStr

            # adding park name
            wcell4 = ws.cell(i, 4)
            wcell4.value = parkName

            # adding park URL
            wcell5 = ws.cell(i, 5)
            wcell5.value = parkURL

            # adding address
            wcell6 = ws.cell(i, 6)
            wcell6.value = streetAndCityStr

            # adding duration
            wcell7 = ws.cell(i, 7)
            wcell7.value = getDuration(streetAndCityStr[0: -6])

            wb.save(filepath)

            driver.get(siteFirst + siteState + "/" + countyStr + "-county")
            time.sleep(1 / 4)
            i += 1

        except:
            print('fail')

driver.close()
